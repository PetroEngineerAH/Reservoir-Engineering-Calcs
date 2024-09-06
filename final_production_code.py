import pandas as pd
import pyodbc 
import os
import openpyxl

import os



cnxn =  pyodbc.connect('Driver={SQL Server};'
        'Server=bogsql01\corex;'
        'Database=OFM_DB;'
        'Trusted_Connections=yes;')
          
script=""" SELECT*FROM PROD_DATA"""
df=pd.read_sql_query(script,cnxn)
data=df[['L_NAME','DATE_TIME','HOURS_ON','gas','OIL','WATER', 'T_PIP']]

names = ['AKIRA', 'TIGANA', 'TIGUI', 'TILO', 'TOTORO', 'BACANO', 'BACANO-OESTE', 'JACANA', 'KITARO', 'SARDANA'] 


data = data[data['L_NAME'].str.contains('|'.join(names))]

print(data['L_NAME'].unique())

data = data.rename(columns = {'L_NAME':'UNIQUEID', 'DATE_TIME':'DATE', 'gas':'GAS', 'T_PIP':'PIP', 'HOURS_ON':'HOURS'})


# # CHANGE THESE TO CHANGE THE DATES: AH
# Date to get data from -> 
recent_update_date = '2024-07-15'
# Set as to yesterday
recent_date = '2024-08-27'
data_recent_date = data[data['DATE']==recent_date]
wells_formation = data_recent_date['UNIQUEID'].unique()

data_specific_wells = data[data['UNIQUEID'].isin(wells_formation)]

data_spec_update = data_specific_wells[data_specific_wells['DATE'] >= recent_update_date]


data = data_spec_update


# CHANGE TO USE DIFFERENT SPREADSHEETS (file path)
spreadsheet = os.listdir("S:/Engineering/Res Engineering/Reservoir Production Surveillance/~Python Code/")

data['UWI'] = data['UNIQUEID']


data['UWI'] = data['UWI'].str.replace("_", " ")

#It goes through each of spreadsheet names and splits that to take well name without production
spreadsheet_names = []
for name in spreadsheet:
    well_name = name.split('Pr')[0][:-1]
    print(well_name)
    spreadsheet_names.append(well_name)
    
# def prepend_to_filename(filename, prefix):
#     # Prepend the prefix to the filename
#     parts = filename.split('.')
#     parts[0] = prefix + parts[0]
#     return '.'.join(parts)

#Function that finds last row with data, so we know where to put data in spreadsheet
def find_last_row_with_numbers(sheet):
    for row in range(sheet.max_row, 0, -1): #start with the max row (last) and go all the way to 0 with increment of -1
        for col in range(1, sheet.max_column + 1): #start first column and goes until last column, iterates through it
            cell_value = sheet.cell(row=row, column=col).value #cell value gets the cell value for the last column
            if isinstance(cell_value, (int, float)): #if the cell value a number, return that row and pastes new data
                return row
    return 0
    
  
sql_wnames = data['UWI'].unique()
for wname in spreadsheet_names:
    if wname == 'Jacana-31' or wname == 'Akira-1' or wname == 'Bacano Oeste-15' or wname == 'Jacana-51':
        continue #Skip code below and go next item in the list
    if "Horizontal" in wname:
        continue
    print(wname)
    # Path to read in spreadsheets
    filename = os.listdir("S:/Engineering/Res Engineering/Reservoir Production Surveillance/~Python Code/" + str(wname) + ' Production/')[-1]
    if filename[:2] == '~$': #If the first two characters of the file name are this, we remove this
        filename = filename[2:]
    sprd = pd.read_excel("S:/Engineering/Res Engineering/Reservoir Production Surveillance/~Python Code/" + str(wname) + ' Production/' + filename, 'Production Summary')
    wname_upper = str(wname).upper()
    # print(wname_upper)
    well_data = data[data['UWI'] == wname_upper]
    well_data = well_data.sort_values(by = 'DATE')
    sprd = sprd.rename(columns = {'Unnamed: 0':'DATE', 'Unnamed: 1':'HOURS', 'Unnamed: 3':'OIL', 'Unnamed: 4':'WATER', 'Unnamed: 5':'GAS', 'Unnamed: 11':'PIP'})
    # File path to export back to spreadsheets
    filename_exp = "S:/Engineering/Res Engineering/Reservoir Production Surveillance/~Python Code/" + str(wname) + ' Production/' + filename
    # if wname == 'Bacano Oeste-15':       #Uncomment if you want only one specific well
    workbook = openpyxl.load_workbook(filename_exp, keep_vba=True)
    sheet = workbook['Production Summary']
    last_data_row = find_last_row_with_numbers(sheet)
    first_empty_row = last_data_row+1
    well_data = well_data[['DATE', 'HOURS', 'GAS', 'OIL', 'WATER', 'PIP']]
    data_vals = well_data.values.tolist()
    columns = [1,2,6,4,5,12]
    for row in data_vals:
        for col_num, value in zip(columns, row):
            sheet.cell(row = first_empty_row, column=col_num, value=value)
        first_empty_row+=1
            
                  
    workbook.save(filename_exp)