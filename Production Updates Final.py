
import pandas as pd
import pyodbc 
import os
import openpyxl

import os



cnxn =  pyodbc.connect('Driver={SQL Server};'
        'Server=bogsql01\corex;'
        'Database=OFM_DB;'
        'Trusted_Connections=yes;')
          
script=""" SELECT*FROM PROD_DATA"""
df=pd.read_sql_query(script,cnxn)
data=df[['L_NAME','DATE_TIME','HOURS_ON','gas','OIL','WATER', 'T_PIP']]
# dfs = {}
# data.to_excel("S:/Users/AHernandez/production_report_proddata.xlsx")
# folder = os.listdir("S:\Engineering\Res Engineering\Reservoir Production Surveillance\~Python Code")
# data = pd.read_excel("S:/Users/AHernandez/production_report_proddata.xlsx") 
# data=data[['UNIQUEID','WELL_ID','DATE','HOURS','GAS','OIL','WATER']]
names = ['AKIRA', 'TIGANA', 'TIGUI', 'TILO', 'TOTORO', 'BACANO', 'BACANO-OESTE', 'JACANA', 'TUA'] 
# # desired_files = []
# # print(data['UNIQUEID'].unique())

data = data[data['L_NAME'].str.contains('|'.join(names))]
# # data_filt.to_excel("S:/Users/Ahernandez/MG_AH_stuff/prod_filtered_wells.xlsx")
# # data = pd.read_excel("S:/Users/Ahernandez/MG_AH_stuff/prod_filtered_wells.xlsx")
print(data['L_NAME'].unique())

data = data.rename(columns = {'L_NAME':'UNIQUEID', 'DATE_TIME':'DATE', 'gas':'GAS', 'T_PIP':'PIP', 'HOURS_ON':'HOURS'})


# # print(data['DATE'].max())
# # CHANGE THESE AH
recent_update_date = '2024-07-15'
recent_date = '2024-08-08'
data_recent_date = data[data['DATE']==recent_date]
wells_formation = data_recent_date['UNIQUEID'].unique()

data_specific_wells = data[data['UNIQUEID'].isin(wells_formation)]

data_spec_update = data_specific_wells[data_specific_wells['DATE'] >= recent_update_date]

# # data_spec_update.to_excel("S:/Users/Ahernandez/MG_AH_stuff/prod_filtered_wells_v2.xlsx")

data = data_spec_update

spreadsheet = os.listdir("S:/Engineering/Res Engineering/Reservoir Production Surveillance/~Python Code/")
# print(data['UNIQUEID'].str.split(":"))
# data["UWI"], data['ZONE']=data['UNIQUEID'].str.split(":", n=1).str
data['UWI'] = data['UNIQUEID']


data['UWI'] = data['UWI'].str.replace("_", " ")
# data['HOURS'] = 24 - data['HOURS']
# # data=data.rename(columns = {'PIP', 'Choke'})

spreadsheet_names = []
for name in spreadsheet:
    well_name = name.split('Pr')[0][:-1]
    print(well_name)
    spreadsheet_names.append(well_name)
    
def prepend_to_filename(filename, prefix):
    # Prepend the prefix to the filename
    parts = filename.split('.')
    parts[0] = prefix + parts[0]
    return '.'.join(parts)


def find_last_row_with_numbers(sheet):
    for row in range(sheet.max_row, 0, -1):
        for col in range(1, sheet.max_column + 1):
            cell_value = sheet.cell(row=row, column=col).value
            if isinstance(cell_value, (int, float)):
                return row
    return 0
    
    
# print(os.listdir("S:/Engineering/Res Engineering/Reservoir Production Surveillance/~Python Code/" + 'AKIRA-24' + ' Production/')[-1])
    
sql_wnames = data['UWI'].unique()
for wname in spreadsheet_names:
    if wname == 'Jacana-31':
        continue
    print(wname)
    if "Horizontal" in wname:
        continue
    filename = os.listdir("S:/Engineering/Res Engineering/Reservoir Production Surveillance/~Python Code/" + str(wname) + ' Production/')[-1]
    if filename[:2] == '~$':
        filename = filename[2:]
    print(filename)
    sprd = pd.read_excel("S:/Engineering/Res Engineering/Reservoir Production Surveillance/~Python Code/" + str(wname) + ' Production/' + filename, 'Production Summary')
    wname_upper = str(wname).upper()
    # print(wname_upper)
    well_data = data[data['UWI'] == wname_upper]
    well_data = well_data.sort_values(by = 'DATE')
    sprd = sprd.rename(columns = {'Unnamed: 0':'DATE', 'Unnamed: 1':'HOURS', 'Unnamed: 3':'OIL', 'Unnamed: 4':'WATER', 'Unnamed: 5':'GAS', 'Unnamed: 11':'PIP'})
    filename_exp = "S:/Engineering/Res Engineering/Reservoir Production Surveillance/~Python Code/" + str(wname) + ' Production/' + filename
    # if wname == 'Bacano Oeste-15':
    workbook = openpyxl.load_workbook(filename_exp, keep_vba=True)
    sheet = workbook['Production Summary']
    last_data_row = find_last_row_with_numbers(sheet)
    first_empty_row = last_data_row+1
    well_data = well_data[['DATE', 'HOURS', 'GAS', 'OIL', 'WATER', 'PIP']]
    data_vals = well_data.values.tolist()
    columns = [1,2,6,4,5,12]
    for row in data_vals:
        for col_num, value in zip(columns, row):
            sheet.cell(row = first_empty_row, column=col_num, value=value)
        first_empty_row+=1
            
        # new_file_path = prepend_to_filename.replace('.xlsm', '.xlsx')
        
            
        workbook.save(filename_exp)
    
    
    
    # full_df = pd.concat([sprd, well_data])
    
    # full_df = full_df.drop(columns = {'Unnamed: 0.2', 'Unnamed: 0.1', 'Unnamed: 0', 'UNIQUEID', 'WELL_ID', 'UWI', 'ZONE'})
    # full_df.to_excel('S:/Users/AHernandez/MG_AH_stuff/test_production/' + str(wname) + '_Production.xlsx')
    
    # print(well_data.shape)
    
    
    
    
    
    
    
    



# prod_summary = data['Production Summary']
# prod_summary_selected = prod_summary[['Unnamed: 0','Unnamed: 3','Unnamed: 4','Unnamed: 16']]
# prod_summary_na = prod_summary_selected.fillna(0)
# prod_summary['Unnamed: 10'] = prod_summary['Unnamed: 10'].astype('str')
# prod_summary = prod_summary.drop([0,1,2,3,4])
# columns_names = prod_summary.iloc[0]
# prod_summary = prod_summary[['Unnamed: 0', 'Unnamed: 1', 'Unnamed: 3', 'Unnamed :4', 'Unnamed :5', 'Unnamed :11']]
# prod_summary = prod_summary.rename(columns = {})
# print(prod_summary.dtypes)

# for f in folder :
#     for wellname in names:
#         if wellname in f: 
#             desired_files.append(f) 
# for files in folder :
#     for wellname in names:
#         prod_summary = data['Production Summary']
# for row in df.iterrows():
#     dfs[row["UNIQUEID"].str.split(":")[0]] = row

# for f in folder :
#     for wellname in names:
#         if wellname in f: 
#             desired_files.append(f) 
# for files in folder :
#     for wellname in names:
#         prod_summary = data['Production Summary']


# prod_summary_selected = prod_summary[['Unnamed: 0','Unnamed: 3','Unnamed: 4','Unnamed: 16']]
# prod_summary_na = prod_summary_selected.fillna(0)
# prod_summary['Unnamed: 10'] = prod_summary['Unnamed: 10'].astype('str')
# prod_summary = prod_summary.drop([0,1,2,3,4])
# print(prod_summary.dtypes)

